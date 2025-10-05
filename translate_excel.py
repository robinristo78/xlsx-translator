import os
from openpyxl import load_workbook
from googletrans import Translator

def translate_file(input_path, output_path, target_lang="en"):
    """Translate all text cells in a single Excel file, handling Japanese and formulas."""
    print(f"Translating: {os.path.basename(input_path)} → {os.path.basename(output_path)}")

    # Open workbook in data_only mode to read cell values, not formulas
    wb = load_workbook(input_path, data_only=True)
    translator = Translator()

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                # Only translate if cell contains a non-empty string
                if isinstance(cell.value, str) and cell.value.strip():
                    try:
                        translation = translator.translate(cell.value, src="auto", dest=target_lang)
                        cell.value = translation.text
                    except Exception as e:
                        print(f"⚠️ Skipped '{cell.value}': {e}")

    wb.save(output_path)
    print(f"✅ Saved: {output_path}\n")


def main():
    input_dir = "input"
    output_dir = "output"
    target_lang = "en"  # default target language

    os.makedirs(output_dir, exist_ok=True)

    # List all .xlsx files in input folder
    files = [f for f in os.listdir(input_dir) if f.endswith(".xlsx")]
    if not files:
        print("⚠️ No .xlsx files found in 'input/' directory.")
        return

    for file in files:
        input_path = os.path.join(input_dir, file)
        output_path = os.path.join(output_dir, f"translated_{file}")
        translate_file(input_path, output_path, target_lang)


if __name__ == "__main__":
    main()
