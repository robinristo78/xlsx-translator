import os
from openpyxl import load_workbook

def apply_translations(original_file, translated_file, output_file):
    """
    Replaces text in the original Excel file with translations
    from a manually translated single-column file.
    """
    # Load original workbook
    wb = load_workbook(original_file)
    # Load translated workbook
    trans_wb = load_workbook(translated_file)
    trans_ws = trans_wb.active

    # Read all translated texts from column A
    translations = [cell.value for cell in trans_ws['A'] if cell.value is not None]

    idx = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    if idx < len(translations):
                        cell.value = translations[idx]
                        idx += 1

    wb.save(output_file)
    print(f"✅ Translated file saved as '{output_file}'")


if __name__ == "__main__":
    os.makedirs("input", exist_ok=True)
    os.makedirs("output", exist_ok=True)

    # Loop through all Excel files in input folder
    for f in os.listdir("input"):
        if f.endswith(".xlsx") and not f.startswith("to_translate_"):
            original_file = os.path.join("input", f)
            translated_file = os.path.join("output", f"to_translate_{f}")  # manually translated
            output_file = os.path.join("output", f"translated_{f}")

            if os.path.exists(translated_file):
                apply_translations(original_file, translated_file, output_file)
            else:
                print(f"⚠️ Translated file not found: {translated_file}")
