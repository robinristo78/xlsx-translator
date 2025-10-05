import os
from openpyxl import load_workbook, Workbook

def export_for_translation(input_file, output_file):
    """
    Reads an Excel file and outputs a clean, single-column file
    with all text strings for manual translation.
    """
    wb = load_workbook(input_file, data_only=True)
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "Translations"

    row_idx = 1
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.strip():
                    # Write text to output
                    out_ws.cell(row=row_idx, column=1, value=cell.value)
                    row_idx += 1

    out_wb.save(output_file)
    print(f"✅ Translation template saved as '{output_file}'")


if __name__ == "__main__":
    os.makedirs("input", exist_ok=True)
    os.makedirs("output", exist_ok=True)

    # Loop through all Excel files in input folder
    files = [f for f in os.listdir("input") if f.endswith(".xlsx")]
    for f in files:
        input_path = os.path.join("input", f)
        output_path = os.path.join("output", f"to_translate_{f}")
        export_for_translation(input_path, output_path)
